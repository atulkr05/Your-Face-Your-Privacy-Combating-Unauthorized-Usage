#!/usr/bin/env python3
import os
import argparse
import logging
from pathlib import Path

import numpy as np
import pandas as pd
from deepface import DeepFace
from scipy.spatial.distance import cosine, euclidean
from scipy.linalg import norm
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

MODELS = ['ArcFace', 'Dlib', 'VGG-Face', 'SFace', 'Facenet', 'Facenet512', 'DeepFace']
DISTANCE_METRICS = ['cosine']
THRESHOLDS = {
    'VGG-Face': {'cosine': 0.68, 'euclidean': 1.17, 'euclidean_l2': 1.17},
    'Facenet': {'cosine': 0.40, 'euclidean': 10, 'euclidean_l2': 0.80},
    'Facenet512': {'cosine': 0.30, 'euclidean': 23.56, 'euclidean_l2': 1.04},
    'ArcFace': {'cosine': 0.68, 'euclidean': 4.15, 'euclidean_l2': 1.13},
    'Dlib': {'cosine': 0.07, 'euclidean': 0.6, 'euclidean_l2': 0.4},
    'SFace': {'cosine': 0.593, 'euclidean': 10.734, 'euclidean_l2': 1.055},
    'DeepFace': {'cosine': 0.23, 'euclidean': 64, 'euclidean_l2': 0.64},
}
DEFAULT_THRESHOLD = {'cosine': 0.40, 'euclidean': 0.55, 'euclidean_l2': 0.75}
IMAGE_EXTS = ('.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.webp')


def calculate_distance(embedding1, embedding2, metric):
    if metric == 'cosine':
        return cosine(embedding1, embedding2)
    if metric == 'euclidean':
        return euclidean(embedding1, embedding2)
    if metric == 'euclidean_l2':
        n1 = embedding1 / norm(embedding1)
        n2 = embedding2 / norm(embedding2)
        return euclidean(n1, n2)
    raise ValueError(f'Unknown distance metric: {metric}')


def get_threshold(model_name, metric):
    return THRESHOLDS.get(model_name, DEFAULT_THRESHOLD).get(metric, 0.4)


def is_image_file(name):
    return name.lower().endswith(IMAGE_EXTS)


def extract_embedding(image_path, model_name, detector_backend='opencv'):
    try:
        reps = DeepFace.represent(
            img_path=str(image_path),
            model_name=model_name,
            enforce_detection=False,
            detector_backend=detector_backend,
        )
        if not reps:
            return None
        return np.asarray(reps[0]['embedding'], dtype=np.float32)
    except Exception as e:
        logging.warning('Embedding failed for %s (%s): %s', image_path, model_name, e)
        return None


def list_subjects(root):
    root = Path(root)
    return sorted([p.name for p in root.iterdir() if p.is_dir()])


def list_images(folder):
    folder = Path(folder)
    if not folder.exists():
        return []
    return sorted([p.name for p in folder.iterdir() if p.is_file() and is_image_file(p.name)])


def load_gallery_embeddings(gallery_base_path, model_name):
    gallery_base_path = Path(gallery_base_path)
    embeddings = {}
    for subject in list_subjects(gallery_base_path):
        subject_dir = gallery_base_path / subject
        imgs = list_images(subject_dir)
        if not imgs:
            continue
        emb = extract_embedding(subject_dir / imgs[0], model_name)
        if emb is not None:
            embeddings[subject] = emb
    return embeddings


def load_probe_embeddings(probe_base_path, model_name):
    probe_base_path = Path(probe_base_path)
    embeddings = {}
    for subject in list_subjects(probe_base_path):
        subject_dir = probe_base_path / subject
        imgs = list_images(subject_dir)
        if not imgs:
            continue
        embeddings[subject] = {}
        for img in imgs:
            emb = extract_embedding(subject_dir / img, model_name)
            if emb is not None:
                embeddings[subject][img] = emb
    return embeddings


def verify_subjects(gallery_base_path, probe_base_path, log_base_dir):
    log_base_dir = Path(log_base_dir)
    log_base_dir.mkdir(parents=True, exist_ok=True)

    results = []
    subjects = list_subjects(gallery_base_path)
    logging.info('Processing %d subjects', len(subjects))

    for model in MODELS:
        logging.info('Model: %s', model)
        gallery_embeddings = load_gallery_embeddings(gallery_base_path, model)
        probe_embeddings = load_probe_embeddings(probe_base_path, model)

        for metric in DISTANCE_METRICS:
            threshold = get_threshold(model, metric)
            matched_log = open(log_base_dir / f'{model}_{metric}_matched.log', 'w', encoding='utf-8')
            nonmatched_log = open(log_base_dir / f'{model}_{metric}_nonmatched.log', 'w', encoding='utf-8')
            try:
                for subject in subjects:
                    if subject not in gallery_embeddings or subject not in probe_embeddings:
                        continue
                    gallery_emb = gallery_embeddings[subject]
                    total = 0
                    correct = 0
                    for probe_img, probe_emb in probe_embeddings[subject].items():
                        dist = calculate_distance(gallery_emb, probe_emb, metric)
                        verified = dist <= threshold
                        total += 1
                        if verified:
                            correct += 1
                        status = 'Matched' if verified else 'Not Matched'
                        entry = (
                            f'Subject: {subject}, Model: {model}, Metric: {metric}, '
                            f'Probe Image: {probe_img}, Status: {status}, '
                            f'Distance: {dist:.6f}, Threshold: {threshold}\n'
                        )
                        (matched_log if verified else nonmatched_log).write(entry)
                    if total:
                        results.append({
                            'model': model,
                            'metric': metric,
                            'subject': subject,
                            'total': total,
                            'correct': correct,
                            'accuracy': correct / total,
                        })
            finally:
                matched_log.close()
                nonmatched_log.close()
    return pd.DataFrame(results)


def save_results_to_excel(df, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Verification Results'

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    data_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ['Model', 'Metric', 'Subject', 'Total', 'Correct', 'Accuracy (%)']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = data_alignment
        c.border = border

    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        values = [row.model, row.metric, row.subject, row.total, row.correct, round(row.accuracy * 100, 2)]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.alignment = data_alignment
            c.border = border

    summary_start = len(df) + 4
    ws.cell(row=summary_start, column=1, value='SUMMARY BY MODEL AND METRIC').font = Font(bold=True, size=12)
    summary_headers = ['Model', 'Metric', 'Total', 'Correct', 'Accuracy (%)']
    for col, h in enumerate(summary_headers, 1):
        c = ws.cell(row=summary_start + 1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = data_alignment
        c.border = border

    grouped = df.groupby(['model', 'metric'], as_index=False).agg({'total': 'sum', 'correct': 'sum'})
    grouped['accuracy'] = np.where(grouped['total'] > 0, grouped['correct'] / grouped['total'] * 100, 0)
    for i, row in enumerate(grouped.itertuples(index=False), summary_start + 2):
        vals = [row.model, row.metric, int(row.total), int(row.correct), round(row.accuracy, 2)]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.alignment = data_alignment
            c.border = border

    widths = {'A': 15, 'B': 15, 'C': 20, 'D': 12, 'E': 12, 'F': 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    wb.save(output_path)


def save_results_to_csv(df, output_path):
    df_out = df.copy()
    df_out['accuracy_pct'] = (df_out['accuracy'] * 100).round(2)
    df_out.to_csv(output_path, index=False)


def main():
    parser = argparse.ArgumentParser(description='Fast face verification using DeepFace embeddings.')
    parser.add_argument('gallery_base_path', help='Root gallery folder')
    parser.add_argument('probe_base_path', help='Root probe folder')
    parser.add_argument('log_base_dir', help='Directory for logs')
    parser.add_argument('--output_excel', default='verification_results.xlsx', help='Excel output file')
    parser.add_argument('--output_csv', default='verification_results.csv', help='CSV output file')
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format='[%(levelname)s] %(message)s')

    df = verify_subjects(args.gallery_base_path, args.probe_base_path, args.log_base_dir)
    if df.empty:
        print('No results generated.')
        return

    save_results_to_csv(df, args.output_csv)
    save_results_to_excel(df, args.output_excel)
    print(df[['model', 'metric', 'subject', 'total', 'correct', 'accuracy']].to_string(index=False))
    print(f'CSV saved to: {args.output_csv}')
    print(f'Excel saved to: {args.output_excel}')


if __name__ == '__main__':
    main()
